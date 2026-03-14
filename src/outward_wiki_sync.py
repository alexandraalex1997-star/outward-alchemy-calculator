from __future__ import annotations

import json
import os
import re
from pathlib import Path
from typing import Dict, Iterable, List

import pandas as pd
import requests
from bs4 import BeautifulSoup, Tag
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_URL = os.getenv("OUTWARD_WIKI_BASE", "https://outward.wiki.gg").rstrip("/")
DATA_DIR = Path(__file__).resolve().parent / "data"
DATA_DIR.mkdir(exist_ok=True)

RECIPE_PAGES = ["Alchemy", "Cooking", "Crafting/Survival"]
GROUP_PAGES = [
    "Bread", "Egg", "Fish", "Meat", "Mushroom", "Ration Ingredient",
    "Vegetable", "Water", "Basic Armor", "Basic Boots", "Basic Helm",
]
DLC_BADGES = {"The Soroboreans", "The Three Brothers", "Definitive Edition", "DLC"}

HEADERS = {
    "User-Agent": "OutwardCraftingHelper/2.0 (+personal local recipe sync)",
    "Accept": "application/json,text/html;q=0.9,*/*;q=0.8",
}


def normalize(text: str) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip()



def slug_key(text: str) -> str:
    return normalize(text).casefold()



def detect_api_endpoint(session: requests.Session) -> str:
    for endpoint in [f"{BASE_URL}/api.php", f"{BASE_URL}/w/api.php"]:
        try:
            r = session.get(endpoint, params={
                "action": "query",
                "meta": "siteinfo",
                "siprop": "general",
                "format": "json",
            }, timeout=30)
            if r.ok and "query" in r.text:
                return endpoint
        except Exception:
            pass
    raise RuntimeError(f"Could not find a working MediaWiki API endpoint on {BASE_URL}")



def fetch_page_html(session: requests.Session, api_endpoint: str, title: str) -> str:
    r = session.get(api_endpoint, params={
        "action": "parse",
        "page": title,
        "prop": "text",
        "format": "json",
        "formatversion": "2",
        "redirects": "1",
    }, timeout=60)
    r.raise_for_status()
    data = r.json()
    if "error" in data:
        raise RuntimeError(f"Could not parse page {title}: {data['error']}")
    return data["parse"]["text"]



def clean_text_tokens(strings: Iterable[str]) -> List[str]:
    out = []
    for s in strings:
        s = normalize(s)
        if not s:
            continue
        if s in DLC_BADGES:
            continue
        if s.lower() in {"result", "ingredients", "station"}:
            continue
        out.append(s)
    return out



def nearest_section_name(table: Tag) -> str:
    node = table
    while node is not None:
        node = node.previous_sibling
        if isinstance(node, Tag) and node.name in {"h2", "h3", "h4"}:
            text = normalize(" ".join(node.stripped_strings)).replace("[edit]", "").strip()
            return text or "Unknown"
    return "Unknown"



def parse_result_cell(cell: Tag) -> tuple[str, int]:
    text = normalize(" ".join(clean_text_tokens(cell.stripped_strings)))
    text = re.sub(r"\[.*?\]", "", text).strip()
    for pattern in [r"^(?P<qty>\d+)\s*x\s*(?P<name>.+)$", r"^(?P<qty>\d+)x(?P<name>.+)$"]:
        m = re.match(pattern, text, flags=re.I)
        if m:
            return normalize(m.group("name")), int(m.group("qty"))
    return text, 1



def parse_ingredient_cell(cell: Tag) -> List[str]:
    bullet_items = [normalize(" ".join(li.stripped_strings)) for li in cell.find_all("li")]
    bullet_items = [x for x in bullet_items if x and x not in DLC_BADGES]
    if bullet_items:
        return bullet_items

    tokens = clean_text_tokens(cell.stripped_strings)
    cleaned = []
    for token in tokens:
        token = re.sub(r"^\d+\s*x\s*", "", token, flags=re.I).strip()
        if token and token not in DLC_BADGES:
            cleaned.append(token)
    return cleaned



def parse_station_cell(cell: Tag) -> str:
    text = normalize(" ".join(clean_text_tokens(cell.stripped_strings)))
    return text or "None"



def parse_recipe_tables(page_title: str, html: str) -> List[dict]:
    soup = BeautifulSoup(html, "html.parser")
    rows_out: List[dict] = []
    counter = 1
    for table in soup.find_all("table"):
        header_cells = table.find_all("th")
        headers = [normalize(" ".join(th.stripped_strings)).casefold() for th in header_cells]
        if not headers:
            continue
        blob = " | ".join(headers)
        if not all(word in blob for word in ["result", "ingredients", "station"]):
            continue
        section = nearest_section_name(table)
        for tr in table.find_all("tr")[1:]:
            cells = tr.find_all(["td", "th"])
            if len(cells) < 3:
                continue
            result_name, result_qty = parse_result_cell(cells[0])
            ingredients = parse_ingredient_cell(cells[1])
            station = parse_station_cell(cells[2])
            if not result_name or not ingredients:
                continue
            rows_out.append({
                "recipe_id": f"{page_title}-{counter}",
                "recipe_page": page_title,
                "section": section,
                "result": result_name,
                "result_qty": result_qty,
                "station": station,
                "ingredients": "|".join(ingredients[:4]),
            })
            counter += 1
    return rows_out



def parse_group_members(html: str, page_title: str) -> List[str]:
    soup = BeautifulSoup(html, "html.parser")
    target_heading = None
    for heading in soup.find_all(["h2", "h3", "h4"]):
        text = normalize(" ".join(heading.stripped_strings)).casefold()
        if "list of ingredients" in text:
            target_heading = heading
            break

    members: List[str] = []
    if target_heading is not None:
        for sib in target_heading.next_siblings:
            if isinstance(sib, Tag) and sib.name in {"h2", "h3", "h4"}:
                break
            if isinstance(sib, Tag):
                for li in sib.find_all("li"):
                    txt = normalize(" ".join(li.stripped_strings))
                    if txt and txt not in DLC_BADGES:
                        members.append(txt)

    if not members:
        for li in soup.find_all("li"):
            txt = normalize(" ".join(li.stripped_strings))
            if not txt or txt in DLC_BADGES:
                continue
            low = txt.casefold()
            if any(x in low for x in ["community", "main page", "all pages", "history", "talk", "register", "sign in", "see also", "categories"]):
                continue
            members.append(txt)

    cleaned = []
    seen = set()
    for item in members:
        item = normalize(re.sub(r"\s*Image:.*$", "", item))
        if not item:
            continue
        k = slug_key(item)
        if k in seen:
            continue
        seen.add(k)
        cleaned.append(item)
    return cleaned



def build_item_index(recipes_df: pd.DataFrame, groups: Dict[str, List[str]]) -> pd.DataFrame:
    produced = recipes_df[["result"]].drop_duplicates().rename(columns={"result": "item"})
    ingredient_rows = []
    for _, row in recipes_df.iterrows():
        for ing in str(row["ingredients"]).split("|"):
            ing = normalize(ing)
            if ing:
                ingredient_rows.append({"item": ing})
    ingredients = pd.DataFrame(ingredient_rows).drop_duplicates() if ingredient_rows else pd.DataFrame(columns=["item"])
    all_items = pd.concat([produced, ingredients], ignore_index=True).drop_duplicates().sort_values("item")
    produced_keys = set(slug_key(x) for x in produced["item"].tolist())
    group_keys = set(groups.keys())
    all_items["is_recipe_output"] = all_items["item"].apply(lambda x: slug_key(x) in produced_keys)
    all_items["is_group_token"] = all_items["item"].apply(lambda x: slug_key(x) in group_keys)
    return all_items.reset_index(drop=True)



def style_header(ws, row: int, labels: List[str], fill_color: str = "1F4E78") -> None:
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin", color="D9D9D9")
    for col_idx, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.fill = fill
        cell.font = font
        cell.border = Border(bottom=thin)
        cell.alignment = Alignment(vertical="center")



def autosize_worksheet(ws, max_width: int = 45) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(val or "")))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_width)



def build_workbook(recipes_df: pd.DataFrame, groups: Dict[str, List[str]], out_path: Path) -> None:
    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "HowTo"
    ws_inv = wb.create_sheet("Inventory")
    ws_rec = wb.create_sheet("Recipes")
    ws_grp = wb.create_sheet("IngredientGroups")
    ws_items = wb.create_sheet("Items")
    ws_dash = wb.create_sheet("Dashboard")

    light_fill = PatternFill("solid", fgColor="D9EAF7")
    accent_fill = PatternFill("solid", fgColor="EAF4E2")
    note_fill = PatternFill("solid", fgColor="FFF2CC")
    bold = Font(bold=True)
    thin = Side(border_style="thin", color="D9D9D9")

    ws_info["A1"] = "Outward Crafting Workbook"
    ws_info["A1"].font = Font(size=15, bold=True)
    ws_info["A3"] = "1) Run outward_wiki_sync.py to refresh recipe data from the wiki"
    ws_info["A4"] = "2) Fill your items into the Inventory sheet"
    ws_info["A5"] = "3) Use app.py for the actual direct + multi-step crafting planner"
    ws_info["A7"] = "Wiki base URL"
    ws_info["B7"] = BASE_URL
    ws_info["A9"] = "Why the app is still important"
    ws_info["A9"].font = bold
    ws_info["A10"] = "The workbook is for storage, filtering, and checking. The app handles generic ingredient groups and multi-step crafting better."
    for cell in ["A3", "A4", "A5"]:
        ws_info[cell].fill = light_fill
    ws_info["A10"].fill = note_fill

    style_header(ws_inv, 1, ["item", "qty", "category", "notes"])
    sample_inventory = [
        ["Wheat", 8, "ingredient", ""],
        ["Clean Water", 4, "ingredient", ""],
        ["Salt", 3, "ingredient", ""],
        ["Egg", 2, "ingredient", ""],
        ["Krimp Nut", 2, "ingredient", ""],
        ["Purpkin", 2, "ingredient", ""],
    ]
    for row_idx, row in enumerate(sample_inventory, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws_inv.cell(row=row_idx, column=col_idx, value=value)
    ws_inv.freeze_panes = "A2"

    rec_headers = ["recipe_id", "recipe_page", "section", "result", "result_qty", "station", "ingredients"]
    style_header(ws_rec, 1, rec_headers)
    for row_idx, record in enumerate(recipes_df.to_dict(orient="records"), start=2):
        for col_idx, h in enumerate(rec_headers, start=1):
            ws_rec.cell(row=row_idx, column=col_idx, value=record[h])
    ws_rec.freeze_panes = "A2"
    ws_rec.auto_filter.ref = f"A1:G{ws_rec.max_row}"

    style_header(ws_grp, 1, ["group_name", "member_item"])
    row = 2
    for group_name, members in sorted(groups.items()):
        for member in members:
            ws_grp.cell(row=row, column=1, value=group_name)
            ws_grp.cell(row=row, column=2, value=member)
            row += 1
    ws_grp.freeze_panes = "A2"
    if ws_grp.max_row > 1:
        ws_grp.auto_filter.ref = f"A1:B{ws_grp.max_row}"

    items_df = build_item_index(recipes_df, groups)
    item_headers = ["item", "is_recipe_output", "is_group_token"]
    style_header(ws_items, 1, item_headers)
    for row_idx, record in enumerate(items_df.to_dict(orient="records"), start=2):
        for col_idx, h in enumerate(item_headers, start=1):
            ws_items.cell(row=row_idx, column=col_idx, value=record[h])
    ws_items.freeze_panes = "A2"
    ws_items.auto_filter.ref = f"A1:C{ws_items.max_row}"

    ws_dash["A1"] = "Quick dashboard"
    ws_dash["A1"].font = Font(size=14, bold=True)
    ws_dash["A3"] = "Inventory rows"
    ws_dash["B3"] = "=COUNTA(Inventory!A:A)-1"
    ws_dash["A4"] = "Known recipes"
    ws_dash["B4"] = "=COUNTA(Recipes!A:A)-1"
    ws_dash["A5"] = "Known ingredient groups"
    ws_dash["B5"] = "=COUNTA(UNIQUE(IngredientGroups!A2:A1048576))"
    ws_dash["A7"] = "Direct-craft logic"
    ws_dash["A7"].font = bold
    ws_dash["A8"] = "Use the Streamlit app for exact craftability. Excel alone is not ideal for recipe groups like Water / Vegetable / Meat."
    for cell in ["A3", "A4", "A5"]:
        ws_dash[cell].fill = accent_fill
    ws_dash["A8"].fill = note_fill

    for ws in [ws_info, ws_inv, ws_rec, ws_grp, ws_items, ws_dash]:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if cell.row == 1:
                    cell.border = Border(bottom=thin)
        autosize_worksheet(ws)

    wb.save(out_path)



def main() -> None:
    session = requests.Session()
    session.headers.update(HEADERS)
    api_endpoint = detect_api_endpoint(session)
    print(f"Using API endpoint: {api_endpoint}")

    all_recipes = []
    for title in RECIPE_PAGES:
        print(f"Fetching recipe page: {title}")
        html = fetch_page_html(session, api_endpoint, title)
        parsed = parse_recipe_tables(title, html)
        print(f"  -> found {len(parsed)} recipe rows")
        all_recipes.extend(parsed)

    if not all_recipes:
        raise RuntimeError("No recipe rows were parsed. Check the wiki structure and parser assumptions.")

    recipes_df = pd.DataFrame(all_recipes)
    recipes_df = recipes_df.drop_duplicates(subset=["result", "station", "ingredients"]).reset_index(drop=True)
    recipes_df.to_csv(DATA_DIR / "recipes.csv", index=False, encoding="utf-8")

    groups: Dict[str, List[str]] = {}
    for title in GROUP_PAGES:
        try:
            print(f"Fetching group page: {title}")
            html = fetch_page_html(session, api_endpoint, title)
            members = parse_group_members(html, title)
            if members:
                groups[slug_key(title)] = members
                print(f"  -> found {len(members)} group members")
        except Exception as exc:
            print(f"  -> warning: failed to parse group page {title}: {exc}")

    (DATA_DIR / "ingredient_groups.json").write_text(json.dumps(groups, ensure_ascii=False, indent=2), encoding="utf-8")
    workbook_path = Path(__file__).resolve().parent / "outward_crafting.xlsx"
    build_workbook(recipes_df, groups, workbook_path)

    print("Done.")
    print(f"Wrote {DATA_DIR / 'recipes.csv'}")
    print(f"Wrote {DATA_DIR / 'ingredient_groups.json'}")
    print(f"Wrote {workbook_path}")


if __name__ == "__main__":
    main()
